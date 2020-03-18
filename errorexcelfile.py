import openpyxl
import mysql.connector
from openpyxl import *
from openpyxl.styles import Font
from openpyxl import load_workbook
from itertools import zip_longest
from mysql.connector.cursor import MySQLCursor
import pandas as pd
import driver
from pandas import ExcelWriter
from pandas import ExcelFile
#source,sourcename both value export file error case

def Sample():
    # SAMPLES 30 VALUES COLLECT CAPTURE IN LISTS
    conn = mysql.connector.connect(user=driver.user, password=driver.password, host=driver.host,database=driver.databasename)
    mycursor = MySQLCursor(conn)

    book = Workbook()
    sheet = book.active

    mycursor.execute('Desc available_trips')
    myresult0 = mycursor.fetchall()  # to get all rows
    res_list = [item for x in zip_longest(*myresult0) for item in x if item]
    finalattrib = res_list[:52]

    for x in range(52):
        c1 = sheet.cell(row=x + 2, column=2)  # attribute name insertion loop
        c1.value = finalattrib[x]

    templist = list(range(1, 31))
    for x in range(30):
        c2 = sheet.cell(row=1, column=5 + x)  # sample string loop
        c2.value = 'Sample' + str(templist[x])

    # print(distinctidd)

    mycursor.execute('SELECT DISTINCT id FROM available_trips ')
    myresult54 = mycursor.fetchall()  # sample 1)id 30 collect from table
    idoutput = [item for x in zip_longest(*myresult54) for item in x if item != -55]
    distinctidd = idoutput[:30]


    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT source FROM available_trips ')
    myresult55 = mycursor.fetchall()  # sample 2)source 30 collect from table
    sourceoutput = [item for x in zip_longest(*myresult55) for item in x if item != -55]
    sourceoutputt = sourceoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT source_name  FROM available_trips ')
    myresult56 = mycursor.fetchall()  # sample 3)sourcename  30 collect from table
    sourcenameoutput = [item for x in zip_longest(*myresult56) for item in x if item != -55]
    sourcenameoutputt = sourcenameoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT destination  FROM available_trips ')
    myresult57 = mycursor.fetchall()  # sample 4)DESTINATION  30 collect from table
    destinationoutput = [item for x in zip_longest(*myresult57) for item in x if item != -55]
    destinationoutputt = destinationoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT destination_name  FROM available_trips ')
    myresult58 = mycursor.fetchall()  # sample 5)destination name  30 collect from table
    destination_nameoutput = [item for x in zip_longest(*myresult58) for item in x if item != -55]
    destination_nameoutputt = destination_nameoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('select distinct travels from available_trips ORDER BY operator ASC ')
    myresult59 = mycursor.fetchall()
    # print(myresult59)                                               #sample 6)travels  30 collect from table
    travelsoutput = [item for x in zip_longest(*myresult59) for item in x if item != -55 and item != '']
    # print(travelsoutput)
    travelsoutputt = travelsoutput[:30]

    #
    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT AC   FROM available_trips ')
    myresult60 = mycursor.fetchall()  # sample 7)AC   30 collect from table
    acoutput = [item for x in zip_longest(*myresult60) for item in x if item != -55]
    acoutputt = acoutput[:30]
    # print(acoutputt)

    mycursor = MySQLCursor(conn)
    mycursor.execute(
        'SELECT distinct arrivalTime FROM available_trips WHERE arrivalTime NOT BETWEEN 0000 AND 2400 order by arrivalTime asc ')
    myresult61 = mycursor.fetchall()
    # print(myresult61)                                                     #sample 8)arrivalTime   30 collect from table
    arrivalTimeoutput = [item for x in zip_longest(*myresult61) for item in x if item != -55]
    arrivalTimeoutputt = arrivalTimeoutput[:30]
    # print(arrivalTimeoutputt)

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT availCatCard  FROM available_trips ')
    myresult62 = mycursor.fetchall()  # sample 9)availCatCard   30 collect from table
    availCatCardoutput = [item for x in zip_longest(*myresult62) for item in x if item != -55]
    availCatCardoutputt = availCatCardoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT availSrCitizen  FROM available_trips ')
    myresult63 = mycursor.fetchall()  # sample 10)availSrCitizen   30 collect from table
    availSrCitizenoutput = [item for x in zip_longest(*myresult63) for item in x if item != -55]
    availSrCitizenoutputt = availSrCitizenoutput[:30]

    # sample 11)availableSeats
    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT availableSeats  FROM available_trips ')
    myresult64 = mycursor.fetchall()  # sample 11)availableSeats   30 collect from table
    availableSeatsoutput = [item for x in zip_longest(*myresult64) for item in x if item != -55]
    availableSeatsoutputt = availableSeatsoutput[:30]
    # print(availableSeatsoutputt)


    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT avlWindowSeats  FROM available_trips ')
    myresult65 = mycursor.fetchall()  # sample12)avlWindowSeats   30 collect from table
    avlWindowSeatsoutput = [item for x in zip_longest(*myresult65) for item in x if item != -55]
    avlWindowSeatsoutputt = avlWindowSeatsoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT boardingTimes  FROM available_trips ')
    myresult66 = mycursor.fetchall()  # sample 13)boardingTimes  30 collect from table
    boardingTimesoutput = [item for x in zip_longest(*myresult66) for item in x if item != -55]
    boardingTimesoutputt = boardingTimesoutput[:30]

    # 14)bookable  30 collect from table

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT bookable  FROM available_trips ')
    myresult67 = mycursor.fetchall()  # sample 14)bookable  30 collect from table
    bookableoutput = [item for x in zip_longest(*myresult67) for item in x if item != -55]
    bookableoutputt = bookableoutput[:30]

#    mycursor = MySQLCursor(conn)
 #   mycursor.execute('SELECT COUNT(bookable) FROM available_trips GROUP BY bookable order by bookable ASC')
  #  myresult67 = mycursor.fetchall()  # sample 14)bookable  30 collect from table
   # bookableoutput = [item for x in zip_longest(*myresult67) for item in x if item != -55]
    #bookableoutputt = bookableoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT bpDpSeatLayout  FROM available_trips ')
    myresult68 = mycursor.fetchall()  # sample          15 bpDpSeatLayout   30 collect from table
    bpDpSeatLayoutoutput = [item for x in zip_longest(*myresult68) for item in x if item != -55]
    bpDpSeatLayoutoutputt = bpDpSeatLayoutoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT busImageCount  FROM available_trips ')
    myresult69 = mycursor.fetchall()
    # print(myresult69)                                       #sample 16busImageCount   30 collect from table
    busImageCountoutput = [item for x in zip_longest(*myresult69) for item in x if item != -55]
    busImageCountoutputt = busImageCountoutput[:30]
    # print(busImageCountoutputt)                                       #sample 16busImageCount   30 collect from table

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT busServiceId  FROM available_trips ')
    myresult70 = mycursor.fetchall()  # sample   17 busServiceId   30 collect from table
    busServiceIdoutput = [item for x in zip_longest(*myresult70) for item in x if item != -55]
    busServiceIdoutputt = busServiceIdoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT busType  FROM available_trips ')
    myresult71 = mycursor.fetchall()  # sample 18busType   30 collect from table
    busTypeoutput = [item for x in zip_longest(*myresult71) for item in x if item != -55]
    busTypeoutputt = busTypeoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT busTypeId  FROM available_trips ')
    myresult72 = mycursor.fetchall()  # sample 19busTypeId  30 collect from table
    busTypeIdoutput = [item for x in zip_longest(*myresult72) for item in x if item != -55]
    busTypeIdoutputt = busTypeIdoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT cancellationPolicy   FROM available_trips ')
    myresult73 = mycursor.fetchall()  # sample 20cancellationPolicy  30 collect from table
    cancellationPolicyoutput = [item for x in zip_longest(*myresult73) for item in x if item != -55]
    cancellationPolicyoutputt = cancellationPolicyoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT departureTime  FROM available_trips ')
    myresult74 = mycursor.fetchall()  # sample 21departureTime  30 collect from table
    departureTimeoutput = [item for x in zip_longest(*myresult74) for item in x if item != -55]
    departureTimeoutputt = departureTimeoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT doj  FROM available_trips ')
    myresult75 = mycursor.fetchall()  # sample 22doj   30 collect from table
    dojoutput = [item for x in zip_longest(*myresult75) for item in x if item != -55]
    dojoutputt = dojoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT dropPointMandatory  FROM available_trips ')
    myresult76 = mycursor.fetchall()  # sample 23dropPointMandatory   30 collect from table
    dropPointMandatoryoutput = [item for x in zip_longest(*myresult76) for item in x if item != -55]
    dropPointMandatoryoutputt = dropPointMandatoryoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT droppingTimes  FROM available_trips ')
    myresult77 = mycursor.fetchall()  # sample 24droppingTimes   30 collect from table
    droppingTimesoutput = [item for x in zip_longest(*myresult77) for item in x if item != -55]
    droppingTimesoutputt = droppingTimesoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT fareDetails  FROM available_trips ')
    myresult78 = mycursor.fetchall()  # sample 25fareDetails   30 collect from table
    fareDetailsoutput = [item for x in zip_longest(*myresult78) for item in x if item != -55]
    fareDetailsoutputt = fareDetailsoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT fares  FROM available_trips ')
    myresult79 = mycursor.fetchall()  # sample 26fares   30 collect from table
    faresoutput = [item for x in zip_longest(*myresult79) for item in x if item != -55]
    faresoutputt = faresoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT flatComApplicable  FROM available_trips ')
    myresult80 = mycursor.fetchall()  # sample 27flatComApplicable  30 collect from table
    flatComApplicableoutput = [item for x in zip_longest(*myresult80) for item in x if item != -55]
    flatComApplicableoutputt = flatComApplicableoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT gdsCommission  FROM available_trips ')
    myresult81 = mycursor.fetchall()  # sample 28gdsCommission   30 collect from table
    gdsCommissionoutput = [item for x in zip_longest(*myresult81) for item in x if item != -55]
    gdsCommissionoutputt = gdsCommissionoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT idProofRequired  FROM available_trips ')
    myresult82 = mycursor.fetchall()  # sample 29idProofRequired  30 collect from table
    idProofRequiredoutput = [item for x in zip_longest(*myresult82) for item in x if item != -55]
    idProofRequiredoutputt = idProofRequiredoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT liveTrackingAvailable  FROM available_trips ')
    myresult83 = mycursor.fetchall()  # sample 30liveTrackingAvailable   30 collect from table
    liveTrackingAvailableoutput = [item for x in zip_longest(*myresult83) for item in x if item != -55]
    liveTrackingAvailableoutputt = liveTrackingAvailableoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT maxSeatsPerTicket  FROM available_trips ')
    myresult84 = mycursor.fetchall()  # sample 31maxSeatsPerTicket   30 collect from table
    maxSeatsPerTicketoutput = [item for x in zip_longest(*myresult84) for item in x if item != -55]
    maxSeatsPerTicketoutputt = maxSeatsPerTicketoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT nonAC  FROM available_trips ')
    myresult85 = mycursor.fetchall()  # sample 32nonAC   30 collect from table
    nonACoutput = [item for x in zip_longest(*myresult85) for item in x if item != -55]
    nonACoutputt = nonACoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('select distinct operator from available_trips ORDER BY operator ASC')
    myresult86 = mycursor.fetchall()  # sample 33operator  30 collect from table
    operatoroutput = [item for x in zip_longest(*myresult86) for item in x if item != -55]
    operatoroutputt = operatoroutput[:30]
    # print(operatoroutputt)

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT otgEnabled  FROM available_trips ')
    myresult87 = mycursor.fetchall()  # sample 34otgEnabled   30 collect from table
    otgEnabledoutput = [item for x in zip_longest(*myresult87) for item in x if item != -55]
    otgEnabledoutputt = otgEnabledoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT otgPolicy  FROM available_trips ')
    myresult88 = mycursor.fetchall()  # sample 35otgPolicy  30 collect from table
    otgPolicyoutput = [item for x in zip_longest(*myresult88) for item in x if item != -55]
    otgPolicyoutputt = otgPolicyoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT partialCancellationAllowed   FROM available_trips ')
    myresult89 = mycursor.fetchall()  # sample 36partialCancellationAllowed  30 collect from table
    partialCancellationAllowedoutput = [item for x in zip_longest(*myresult89) for item in x if item != -55]
    partialCancellationAllowedoutputt = partialCancellationAllowedoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT partnerBaseCommission  FROM available_trips ')
    myresult90 = mycursor.fetchall()  # sample 37partnerBaseCommission  30 collect from table
    partnerBaseCommissionoutput = [item for x in zip_longest(*myresult90) for item in x if item != -55]
    partnerBaseCommissionoutputt = partnerBaseCommissionoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT primaryPaxCancellable  FROM available_trips ')
    myresult91 = mycursor.fetchall()  # sample 38primaryPaxCancellable   30 collect from table
    primaryPaxCancellableoutput = [item for x in zip_longest(*myresult91) for item in x if item != -55]
    primaryPaxCancellableoutputt = primaryPaxCancellableoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT routeId  FROM available_trips ')
    myresult92 = mycursor.fetchall()  # sample 39routeId  30 collect from table
    routeIdoutput = [item for x in zip_longest(*myresult92) for item in x if item != -55]
    routeIdoutputt = routeIdoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT rtc  FROM available_trips ')
    myresult93 = mycursor.fetchall()  # sample 40rtc   30 collect from table
    rtcoutput = [item for x in zip_longest(*myresult93) for item in x if item != -55]
    rtcoutputt = rtcoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT seater  FROM available_trips ')
    myresult94 = mycursor.fetchall()  # sample 41seater  30 collect from table
    seateroutput = [item for x in zip_longest(*myresult94) for item in x if item != -55]
    seateroutputt = seateroutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT selfInventory  FROM available_trips ')
    myresult95 = mycursor.fetchall()  # sample 42selfInventory   30 collect from table
    selfInventoryoutput = [item for x in zip_longest(*myresult95) for item in x if item != -55]
    selfInventoryoutputt = selfInventoryoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT singleLadies  FROM available_trips ')
    myresult96 = mycursor.fetchall()  # sample 43singleLadies   30 collect from table
    singleLadiesoutput = [item for x in zip_longest(*myresult96) for item in x if item != -55]
    singleLadiesoutputt = singleLadiesoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT sleeper  FROM available_trips ')
    myresult97 = mycursor.fetchall()  # sample 44sleeper   30 collect from table
    sleeperoutput = [item for x in zip_longest(*myresult97) for item in x if item != -55]
    sleeperoutputt = sleeperoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT tatkalTime  FROM available_trips ')
    myresult98 = mycursor.fetchall()  # sample 45tatkalTime   30 collect from table
    tatkalTimeoutput = [item for x in zip_longest(*myresult98) for item in x if item != -55]
    tatkalTimeoutputt = tatkalTimeoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT vehicleType  FROM available_trips ')
    myresult99 = mycursor.fetchall()  # sample 46vehicleType   30 collect from table
    vehicleTypeoutput = [item for x in zip_longest(*myresult99) for item in x if item != -55]
    vehicleTypeoutputt = vehicleTypeoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT viaRoutes FROM available_trips ')
    myresult100 = mycursor.fetchall()  # sample 47viaRoutes  30 collect from table
    viaRoutesoutput = [item for x in zip_longest(*myresult100) for item in x if item != -55]
    viaRoutesoutputt = viaRoutesoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT zeroCancellationTime  FROM available_trips ')
    myresult104 = mycursor.fetchall()  # sample 48zeroCancellationTime  30 collect from table
    zeroCancellationTimeoutput = [item for x in zip_longest(*myresult104) for item in x if item != -55]
    zeroCancellationTimeoutputt = zeroCancellationTimeoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT mTicketEnabled  FROM available_trips ')
    myresult105 = mycursor.fetchall()  # sample 49mTicketEnabled   30 collect from table
    mTicketEnabledoutput = [item for x in zip_longest(*myresult105) for item in x if item != -55]
    mTicketEnabledoutputt = mTicketEnabledoutput[:30]

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT sd_id  FROM available_trips ')
    myresult105 = mycursor.fetchall()  # sample 50sd_id  30 collect from table
    sd_idoutput = [item for x in zip_longest(*myresult105) for item in x if item != -55]
    sd_idoutputt = sd_idoutput[:30]

    #51 createDt
    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT   createDt  FROM available_trips ')
    myresult107 = mycursor.fetchall()  # sample 51created_date  30 collect from table
    createDtoutput = [item for x in zip_longest(*myresult107) for item in x if item != -55]
    createDtoutputt = createDtoutput[:30]
    

    #52 created_date
    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT DISTINCT created_date  FROM available_trips ')
    myresult107 = mycursor.fetchall()  # sample 51created_date  30 collect from table
    created_dateoutput = [item for x in zip_longest(*myresult107) for item in x if item != -55]
    created_dateoutputt = created_dateoutput[:30]

    len1 = len(distinctidd)
    for x in range(len1):
        c4 = sheet.cell(row=2, column=5 + x)  # distinct id 30 samples loop
        c4.value = distinctidd[x]

    len2 = len(sourceoutputt)
    for x in range(len2):
        c4 = sheet.cell(row=3, column=5 + x)  # distinct source 30 samples loop
        c4.value = sourceoutputt[x]

    len3 = len(sourcenameoutputt)
    for x in range(len3):
        c4 = sheet.cell(row=4, column=5 + x)  # distinct source 30 samples loop
        c4.value = sourcenameoutputt[x]

    len4 = len(destinationoutputt)
    for x in range(len4):
        c4 = sheet.cell(row=5, column=5 + x)  # distinct destination 30 samples loop
        c4.value = destinationoutputt[x]

    len5 = len(destination_nameoutputt)
    for x in range(len5):
        c4 = sheet.cell(row=6, column=5 + x)  # distinct destination 30 samples loop
        c4.value = destination_nameoutputt[x]

    len6 = len(travelsoutputt)
    for x in range(len6):
        c4 = sheet.cell(row=7, column=5 + x)  # distinct destination 30 samples loop
        c4.value = travelsoutputt[x]

    len7 = len(acoutput)
    for x in range(len7):
        c4 = sheet.cell(row=8, column=5 + x)  # distinct destination 30 samples loop
        c4.value = acoutputt[x]

    len8 = len(arrivalTimeoutputt)
    for x in range(len8):
        c4 = sheet.cell(row=9, column=5 + x)  # distinct destination 30 samples loop
        c4.value = arrivalTimeoutputt[x]

    len9 = len(availCatCardoutputt)
    for x in range(len9):
        c4 = sheet.cell(row=10, column=5 + x)  # distinct destination 30 samples loop
        c4.value = availCatCardoutputt[x]

    len10 = len(availSrCitizenoutputt)
    for x in range(len10):
        c4 = sheet.cell(row=11, column=5 + x)  # distinct destination 30 samples loop
        c4.value = availSrCitizenoutputt[x]

    len11 = len(availableSeatsoutputt)
    for x in range(len11):
        c4 = sheet.cell(row=12, column=5 + x)  # distinct destination 30 samples loop
        c4.value = availableSeatsoutputt[x]

    len12 = len(avlWindowSeatsoutputt)
    for x in range(len12):
        c4 = sheet.cell(row=13, column=5 + x)  # distinct destination 30 samples loop
        c4.value = avlWindowSeatsoutputt[x]

    len13 = len(boardingTimesoutputt)
    for x in range(len13):
        c4 = sheet.cell(row=14, column=5 + x)  # distinct destination 30 samples loop
        c4.value = boardingTimesoutputt[x]

    len14 = len(bookableoutputt)
    for x in range(len14):
        c4 = sheet.cell(row=15, column=5 + x)  # distinct destination 30 samples loop
        c4.value = bookableoutputt[x]

    len15 = len(bpDpSeatLayoutoutputt)
    for x in range(len15):
        c4 = sheet.cell(row=16, column=5 + x)  # distinct destination 30 samples loop
        c4.value = bpDpSeatLayoutoutputt[x]

    len16 = len(busImageCountoutputt)
    for x in range(len16):
        c4 = sheet.cell(row=17, column=5 + x)  # distinct destination 30 samples loop
        c4.value = busImageCountoutputt[x]

    len17 = len(busServiceIdoutputt)
    for x in range(len17):
        c4 = sheet.cell(row=18, column=5 + x)  # distinct destination 30 samples loop
        c4.value = busServiceIdoutputt[x]

    len18 = len(busTypeoutputt)
    for x in range(len18):
        c4 = sheet.cell(row=19, column=5 + x)  # distinct destination 30 samples loop
        c4.value = busTypeoutputt[x]
    # sourceoutputt[x]

    len19 = len(busTypeIdoutputt)
    for x in range(len19):
        c4 = sheet.cell(row=20, column=5 + x)  # distinct destination 30 samples loop
        c4.value = busTypeIdoutputt[x]

    len20 = len(cancellationPolicyoutputt)
    for x in range(len20):
        c4 = sheet.cell(row=21, column=5 + x)  # distinct destination 30 samples loop
        c4.value = cancellationPolicyoutputt[x]

    len21 = len(departureTimeoutputt)
    for x in range(len21):
        c4 = sheet.cell(row=22, column=5 + x)  # distinct destination 30 samples loop
        c4.value = departureTimeoutputt[x]

    len22 = len(dojoutputt)
    for x in range(len22):
        c4 = sheet.cell(row=23, column=5 + x)  # distinct destination 30 samples loop
        c4.value = dojoutputt[x]

    len23 = len(dropPointMandatoryoutputt)
    for x in range(len23):
        c4 = sheet.cell(row=24, column=5 + x)  # distinct destination 30 samples loop
        c4.value = dropPointMandatoryoutputt[x]

    len24 = len(droppingTimesoutputt)
    for x in range(len24):
        c4 = sheet.cell(row=25, column=5 + x)  # distinct destination 30 samples loop
        c4.value = droppingTimesoutput[x]

    len25 = len(fareDetailsoutputt)
    for x in range(len25):
        c4 = sheet.cell(row=26, column=5 + x)  # distinct destination 30 samples loop
        c4.value = fareDetailsoutputt[x]

    len26 = len(faresoutputt)
    for x in range(len26):
        c4 = sheet.cell(row=27, column=5 + x)  # distinct destination 30 samples loop
        c4.value = faresoutputt[x]

    len27 = len(flatComApplicableoutputt)
    for x in range(len27):
        c4 = sheet.cell(row=28, column=5 + x)  # distinct destination 30 samples loop
        c4.value = flatComApplicableoutputt[x]

    len28 = len(gdsCommissionoutputt)
    for x in range(len28):
        c4 = sheet.cell(row=29, column=5 + x)  # distinct destination 30 samples loop
        c4.value = gdsCommissionoutputt[x]

    len29 = len(idProofRequiredoutputt)
    for x in range(len29):
        c4 = sheet.cell(row=30, column=5 + x)  # distinct destination 30 samples loop
        c4.value = idProofRequiredoutputt[x]

    len30 = len(liveTrackingAvailableoutputt)
    for x in range(len30):
        c4 = sheet.cell(row=31, column=5 + x)  # distinct destination 30 samples loop
        c4.value = liveTrackingAvailableoutputt[x]

    len31 = len(maxSeatsPerTicketoutputt)
    for x in range(len31):
        c4 = sheet.cell(row=32, column=5 + x)  # distinct destination 30 samples loop
        c4.value = maxSeatsPerTicketoutputt[x]

    len32 = len(nonACoutputt)
    for x in range(len32):
        c4 = sheet.cell(row=33, column=5 + x)  # distinct destination 30 samples loop
        c4.value = nonACoutputt[x]

    len33 = len(operatoroutputt)
    for x in range(len33):
        c4 = sheet.cell(row=34, column=5 + x)  # distinct destination 30 samples loop
        c4.value = operatoroutputt[x]

    len34 = len(otgEnabledoutputt)
    for x in range(len34):
        c4 = sheet.cell(row=35, column=5 + x)  # distinct destination 30 samples loop
        c4.value = otgEnabledoutputt[x]

    len35 = len(otgPolicyoutputt)
    for x in range(len35):
        c4 = sheet.cell(row=36, column=5 + x)  # distinct destination 30 samples loop
        c4.value = otgPolicyoutputt[x]

    len36 = len(partialCancellationAllowedoutputt)
    for x in range(len36):
        c4 = sheet.cell(row=37, column=5 + x)  # distinct destination 30 samples loop
        c4.value = partialCancellationAllowedoutputt[x]

    len37 = len(partnerBaseCommissionoutputt)
    for x in range(len37):
        c4 = sheet.cell(row=38, column=5 + x)  # distinct destination 30 samples loop
        c4.value = partnerBaseCommissionoutputt[x]

    len38 = len(primaryPaxCancellableoutputt)
    for x in range(len38):
        c4 = sheet.cell(row=39, column=5 + x)  # distinct destination 30 samples loop
        c4.value = primaryPaxCancellableoutputt[x]
    len39 = len(routeIdoutputt)
    for x in range(len39):
        c4 = sheet.cell(row=30, column=5 + x)  # distinct destination 30 samples loop
        c4.value = routeIdoutputt[x]

    len40 = len(rtcoutputt)
    for x in range(len40):
        c4 = sheet.cell(row=41, column=5 + x)  # distinct destination 30 samples loop
        c4.value = rtcoutputt[x]

    len41 = len(seateroutputt)
    for x in range(len41):
        c4 = sheet.cell(row=42, column=5 + x)  # distinct destination 30 samples loop
        c4.value = seateroutputt[x]

    len42 = len(selfInventoryoutputt)
    for x in range(len42):
        c4 = sheet.cell(row=43, column=5 + x)  # distinct destination 30 samples loop
        c4.value = selfInventoryoutputt[x]

    len43 = len(singleLadiesoutputt)
    for x in range(len43):
        c4 = sheet.cell(row=44, column=5 + x)  # distinct destination 30 samples loop
        c4.value = singleLadiesoutputt[x]

    len44 = len(sleeperoutputt)
    for x in range(len44):
        c4 = sheet.cell(row=45, column=5 + x)  # distinct destination 30 samples loop
        c4.value = sleeperoutputt[x]

    len45 = len(tatkalTimeoutputt)
    for x in range(len45):
        c4 = sheet.cell(row=46, column=5 + x)  # distinct destination 30 samples loop
        c4.value = tatkalTimeoutputt[x]

    len46 = len(vehicleTypeoutputt)
    for x in range(len46):
        c4 = sheet.cell(row=47, column=5 + x)  # distinct destination 30 samples loop
        c4.value = vehicleTypeoutputt[x]

    len47 = len(viaRoutesoutputt)
    for x in range(len47):
        c4 = sheet.cell(row=48, column=5 + x)  # distinct destination 30 samples loop
        c4.value = viaRoutesoutputt[x]

    len48 = len(zeroCancellationTimeoutputt)
    for x in range(len48):
        c4 = sheet.cell(row=49, column=5 + x)  # distinct destination 30 samples loop
        c4.value = zeroCancellationTimeoutputt[x]

    len49 = len(mTicketEnabledoutputt)
    for x in range(len49):
        c4 = sheet.cell(row=50, column=5 + x)  # distinct destination 30 samples loop
        c4.value = mTicketEnabledoutputt[x]

    len50 = len(sd_idoutputt)
    for x in range(len50):
        c4 = sheet.cell(row=51, column=5 + x)  # distinct destination 30 samples loop
        c4.value = sd_idoutputt[x]


    len51 = len(createDtoutputt)
    for x in range(len51):
        c4 = sheet.cell(row=52, column=5 + x)  # distinct destination 30 samples loop
        c4.value = createDtoutputt[x]    

    len52 = len(created_dateoutputt)
    for x in range(len52):
        c4 = sheet.cell(row=53, column=5 + x)  # distinct destination 30 samples loop
        c4.value = created_dateoutputt[x]


    for x in range(34):
        sheet.cell(row=1, column=1 + x).font = Font(size=15)

    for x in range(52):
        sheet.cell(row=2 + x, column=2).font = Font(size=15)

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 45
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 30
    sheet.column_dimensions['K'].width = 30
    sheet.column_dimensions['L'].width = 30
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 30
    sheet.column_dimensions['O'].width = 30
    sheet.column_dimensions['P'].width = 30
    sheet.column_dimensions['Q'].width = 30
    sheet.column_dimensions['R'].width = 30
    sheet.column_dimensions['S'].width = 30
    sheet.column_dimensions['T'].width = 30
    sheet.column_dimensions['U'].width = 30
    sheet.column_dimensions['V'].width = 30
    sheet.column_dimensions['W'].width = 30
    sheet.column_dimensions['X'].width = 30
    sheet.column_dimensions['Y'].width = 30
    sheet.column_dimensions['Z'].width = 30
    sheet.column_dimensions['AA'].width = 30
    sheet.column_dimensions['AB'].width = 30
    sheet.column_dimensions['AC'].width = 30
    sheet.column_dimensions['AD'].width = 30
    sheet.column_dimensions['AE'].width = 30
    sheet.column_dimensions['AF'].width = 30
    sheet.column_dimensions['AG'].width = 30
    sheet.column_dimensions['AH'].width = 30
    sheet.column_dimensions['AI'].width = 30
    sheet.column_dimensions['AJ'].width = 30
    sheet.column_dimensions['AK'].width = 30


    numberlist = list(range(1, 53))
    for x in range(52):  # attribute number insertion loop
        c3 = sheet.cell(row=2 +x , column=1)
        c3.value = numberlist[x]
    sheet['A1'] = "Attribute Number"
    sheet['B1'] = "Attribute Name"
    sheet['C1'] = "Min"
    sheet['D1'] = "Max"


    #min max insertion logic
    path = driver.desktop+'ParentMainAnalysis.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    minlist=[]
    maxlist=[]
    
    count=0
    for x in range(52):
        cell_obj = sheet_obj.cell(row=8+x, column=4)
        count = cell_obj.value
        if count =='' or count==None:
            count ='BLANK'
        minlist.insert(x, count)
    
    count=0
    for x in range(52):
        cell_obj = sheet_obj.cell(row=8+x, column=5)
        count = cell_obj.value
        maxlist.insert(x, count)

    for x in range(52):  # min values insertion loop
        c3 = sheet.cell(row=2 + x, column=3)
        c3.value = minlist[x]

    for x in range(52):  # max values insertion loop
        c3 = sheet.cell(row=2 + x, column=4)
        c3.value = maxlist[x]

  #  print(minlist)
   # print(maxlist)



    book.save(driver.desktop+'Sample30.xlsx')


    conn.close()


def Sourceerror():  # this function exports source,sourcename error file
    conn = mysql.connector.connect(user=driver.user, password=driver.password, host=driver.host,database=driver.databasename)
    book = Workbook()
    sheet1 = book.active

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT distinct source,source_name FROM available_trips ')
    sbothval = mycursor.fetchall()

    # dictionary conversion & dataframe assignment
    sbothvall = dict(sbothval)
    pd.DataFrame(sbothvall.items())
    svaldf = pd.DataFrame(sbothvall.items(), columns=['source', 'source_name'])

    # case1
    # captures repeat value according to 'source name' column but print both columns
    duplicate_bool = svaldf.duplicated(subset=['source_name'], keep=False)
    sduplicate = svaldf.loc[duplicate_bool == True]
    sduplicate1 = sduplicate
   #print(sduplicate1.shape[0])  # shape returns number of rows

    # case1
    # captures repeat value according to 'source' column but print both columns
    duplicate_bool = svaldf.duplicated(subset=['source'], keep=False)
    sduplicate = svaldf.loc[duplicate_bool == True]
    sduplicate2 = sduplicate
  # print(sduplicate2.shape[0])

    if sduplicate1.shape[0]:
        sduplicate = sduplicate1
    else:
        sduplicate = sduplicate2

    # insertion loop of values to excel file
    s = sduplicate.get(["source"])
    slist = []  # empty list
    for row in s.itertuples():
        mylist = [row.source]
        slist.append(mylist)
    sfval = [item for x in zip_longest(*slist) for item in x if item != -55]

    sname = sduplicate.get(["source_name"])
    snamelist = []
    for row in sname.itertuples():
        mylist = [row.source_name]
        snamelist.append(mylist)
    snamefval = [item for x in zip_longest(*snamelist) for item in x if item != -55]

    # writeback code to excel file
    sheet1.row_dimensions[1].height = 25
    sheet1['C1'] = "ERROR CASE:THIS SOURCENAMES DOESNOT HAVE UNIQUE 'SOURCEid'"
    sheet1.column_dimensions['C'].width = 85
    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 30
    sheet1.cell(row=2, column=1).font = Font(size=15)
    sheet1.cell(row=2, column=2).font = Font(size=15)
    sheet1.cell(row=1, column=3).font = Font(size=15)
    sheet1['A2'] = "source"
    sheet1['B2'] = "source_name"

    for x in range(len(sfval)):
        c1 = sheet1.cell(row=x + 3, column=1)  # source  insertion loop
        c1.value = sfval[x]

    for x in range(len(snamefval)):
        c1 = sheet1.cell(row=x + 3, column=2)  # source name insertion loop
        c1.value = snamefval[x]

    book.save(driver.desktop+'error\\ERRORsourceandsourcename.xlsx')
    conn.close()


def Destinationerror():  # this function exports destination,destination_name error file
    conn = mysql.connector.connect(user=driver.user, password=driver.password, host=driver.host,database=driver.databasename)
    book = Workbook()
    sheet1 = book.active

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT distinct destination,destination_name FROM available_trips ')
    sbothval = mycursor.fetchall()

    # dictionary conversion & dataframe assignment
    sbothvall = dict(sbothval)
    pd.DataFrame(sbothvall.items())
    svaldf = pd.DataFrame(sbothvall.items(), columns=['destination', 'destination_name'])

    # case1
    # captures repeat value according to 'destination name' column but print both columns
    duplicate_bool = svaldf.duplicated(subset=['destination_name'], keep=False)
    sduplicate = svaldf.loc[duplicate_bool == True]
    sduplicate1 = sduplicate
 #  print(sduplicate1.shape[0])  # shape returns number of rows

    # case1
    # captures repeat value according to 'destination' column but print both columns
    duplicate_bool = svaldf.duplicated(subset=['destination'], keep=False)
    sduplicate = svaldf.loc[duplicate_bool == True]
    sduplicate2 = sduplicate
 #  print(sduplicate2.shape[0])

    if sduplicate1.shape[0]:
        sduplicate = sduplicate1
    else:
        sduplicate = sduplicate2

    # insertion loop of values to excel file
    s = sduplicate.get(["destination"])
    slist = []  # empty list
    for row in s.itertuples():
        mylist = [row.destination]
        slist.append(mylist)
    sfval = [item for x in zip_longest(*slist) for item in x if item != -55]

    sname = sduplicate.get(["destination_name"])
    snamelist = []
    for row in sname.itertuples():
        mylist = [row.destination_name]
        snamelist.append(mylist)
    snamefval = [item for x in zip_longest(*snamelist) for item in x if item != -55]

    # writeback code to excel file
    sheet1.row_dimensions[1].height = 25
    sheet1['C1'] = "ERROR CASE:THIS destination names DOESNOT HAVE UNIQUE 'destinationid'"
    sheet1.column_dimensions['C'].width = 100
    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 30
    sheet1.cell(row=2, column=1).font = Font(size=15)
    sheet1.cell(row=2, column=2).font = Font(size=15)
    sheet1.cell(row=1, column=3).font = Font(size=15)
    sheet1['A2'] = "destination"
    sheet1['B2'] = "destination_name"

    for x in range(len(sfval)):
        c1 = sheet1.cell(row=x + 3, column=1)  # destination  insertion loop
        c1.value = sfval[x]

    for x in range(len(snamefval)):
        c1 = sheet1.cell(row=x + 3, column=2)  # destination name insertion loop
        c1.value = snamefval[x]

    book.save(driver.desktop+'error\\ERRORdestANDdestname.xlsx')
    conn.close()
    return 11


def Travelserror():  # this function exports travels,operator error file
    conn = mysql.connector.connect(user=driver.user, password=driver.password, host=driver.host,database=driver.databasename)
    book = Workbook()
    sheet1 = book.active

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT distinct travels,operator FROM available_trips ')
    sbothval = mycursor.fetchall()

    # dictionary conversion & dataframe assignment
    sbothvall = dict(sbothval)
    pd.DataFrame(sbothvall.items())
    svaldf = pd.DataFrame(sbothvall.items(), columns=['travels', 'operator'])

    # case1
    # captures repeat value according to 'travels name' column but print both columns
    duplicate_bool = svaldf.duplicated(subset=['operator'], keep=False)
    sduplicate = svaldf.loc[duplicate_bool == True]
    sduplicate1 = sduplicate
 #  print(sduplicate1.shape[0])  # shape returns number of rows

    # case1
    # captures repeat value according to 'travels' column but print both columns
    duplicate_bool = svaldf.duplicated(subset=['travels'], keep=False)
    sduplicate = svaldf.loc[duplicate_bool == True]
    sduplicate2 = sduplicate
#   print(sduplicate2.shape[0])

    if sduplicate1.shape[0]:
        sduplicate = sduplicate1
    else:
        sduplicate = sduplicate2

    # insertion loop of values to excel file
    s = sduplicate.get(["travels"])
    slist = []  # empty list
    for row in s.itertuples():
        mylist = [row.travels]
        slist.append(mylist)
    sfval = [item for x in zip_longest(*slist) for item in x if item != -55]

    sname = sduplicate.get(["operator"])
    snamelist = []
    for row in sname.itertuples():
        mylist = [row.operator]
        snamelist.append(mylist)
    snamefval = [item for x in zip_longest(*snamelist) for item in x if item != -55]

    # writeback code to excel file
    sheet1.row_dimensions[1].height = 25
    sheet1['C1'] = "ERROR CASE:THIS travels names DOESNOT HAVE UNIQUE 'operator'"
    sheet1.column_dimensions['C'].width = 100
    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 30
    sheet1.cell(row=2, column=1).font = Font(size=15)
    sheet1.cell(row=2, column=2).font = Font(size=15)
    sheet1.cell(row=1, column=3).font = Font(size=15)
    sheet1['A2'] = "travels"
    sheet1['B2'] = "operator"

    for x in range(len(sfval)):
        c1 = sheet1.cell(row=x + 3, column=1)  # travels  insertion loop
        c1.value = sfval[x]

    for x in range(len(snamefval)):
        c1 = sheet1.cell(row=x + 3, column=2)  # travels name insertion loop
        c1.value = snamefval[x]

    book.save(driver.desktop+'error\\ERRORtravelsANDoperator.xlsx')
    conn.close()


'''
def output1():
    conn = mysql.connector.connect(user='viplav', password='password', host='127.0.0.1', database='indiabus8')
    book = Workbook()
    sheet1 = book.active

    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT distinct busServiceId,count(tatkalTime),count(routeId),count(operator) FROM available_trips GROUP BY busServiceId ASC ')
    myresult = mycursor.fetchall()
    print(myresult)
   # print(len(myresult
'''

