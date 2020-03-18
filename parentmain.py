from openpyxl import Workbook
import mysql.connector
from openpyxl.styles import Font

from mysql.connector.cursor import MySQLCursor
from openpyxl.styles import PatternFill
import driver
from itertools import zip_longest




def Myfunc():
    conn = mysql.connector.connect(user=driver.user, password=driver.password, host=driver.host,database=driver.databasename)

    book = Workbook()
    sheet= book.active

    mycursor = conn.cursor()
    mycursor.execute('show tables')  # to catch table name
    myresultw = mycursor.fetchall()
    tablename = [item for x in zip_longest(*myresultw) for item in x if item]

    mycursor = conn.cursor()
    mycursor.execute('SELECT id from available_trips')
    myresultq = mycursor.fetchall()  # for total rows output
    totalrows = [item for x in zip_longest(*myresultq) for item in x if item != -55]

    countnew=0
    for x in range(len(totalrows)):
        countnew = countnew + 1


    totalrows[0]= countnew


    mycursor = conn.cursor()
    mycursor.execute('SELECT DATABASE() FROM DUAL')  # to catch db name
    myresultdb = mycursor.fetchall()
    dbname = [item for x in zip_longest(*myresultdb) for item in x if item]

    mycursor = conn.cursor()
    mycursor.execute('SELECT doj FROM available_trips')   #to get blank values as per doj
    myresult51 = mycursor.fetchall()
    dojcountwithblank = [item for x in zip_longest(*myresult51) for item in x if item != -55]
#print(dojcountwithblank)

    countvarblank=0
    countvarnotblank=0
    for x in range(len(dojcountwithblank)):
        if dojcountwithblank[x] == '':
            countvarblank+=1
        else:
            countvarnotblank+=1

  #  print(countvarblank)


    sheet['C4'] = countvarblank
    sheet['C5'] = totalrows[0]-countvarblank

    sheet['C1'] = dbname[0]
    sheet['C2'] = tablename[0]
    sheet['C3'] = totalrows[0]
   # print(totalrows)



    # attribute1:tripid
    mycursor = MySQLCursor(conn)
    mycursor.execute('SELECT COUNT(DISTINCT id) FROM available_trips')  # count id distinct
    myresult1 = mycursor.fetchall()
    idcount = list(myresult1)
    idcountt = idcount[0][0]

    mycursor.execute('SELECT MIN(id) FROM available_trips')  # Min id value
    myresult1 = mycursor.fetchall()
    minid = list(myresult1)
    minidd = minid[0][0]

    if minidd =='':
        minidd='BLANK'

    mycursor.execute('SELECT MAX(id) FROM available_trips')
    myresult1 = mycursor.fetchall()  # max id value
    maxid = list(myresult1)
    maxidd = maxid[0][0]

    mycursor.execute('Desc available_trips')
    myresult0 = mycursor.fetchall()  # to get all rows
    res_list = [item for x in zip_longest(*myresult0) for item in x if item]
    finalattrib = res_list[:52]

    for x in range(52):
        c1 = sheet.cell(row=x + 8, column=2)  # attribute name insertion loop
        c1.value = finalattrib[x]

    sheet['D8'] = minidd
    sheet['E8'] = maxidd


    #book.save("C:\\Users\\AdminPC\\Desktop\\OUTPUTFOLDER\\ParentMainAnalysis.xlsx")

    #return idcountt

    # attribute2:source

    mycursor.execute('SELECT COUNT(DISTINCT source) FROM available_trips')
    myresult2 = mycursor.fetchall()
    sourcecount = list(myresult2)
    sourcecountt = sourcecount[0][0]

    mycursor.execute('SELECT MIN(source) FROM available_trips')
    myresult2 = mycursor.fetchall()
    minsource = list(myresult2)
    minsourcee = minsource[0][0]

    if minsourcee =='':
        minsourcee='BLANK'

    mycursor.execute('SELECT MAX(source) FROM available_trips')
    myresult2 = mycursor.fetchall()
    maxsource = list(myresult2)
    maxsourcee = maxsource[0][0]

    sheet['D9'] = minsourcee
    sheet['E9'] = maxsourcee



    #return sourcecountt


    # attribute3:source_name
    mycursor.execute('SELECT COUNT(DISTINCT source_name) FROM available_trips')
    myresult3 = mycursor.fetchall()
    sourcenamecount = list(myresult3)
    sourcenamecountt = sourcenamecount[0][0]

    mycursor.execute('SELECT MIN(source_name) FROM available_trips')  # Min id value
    myresult1 = mycursor.fetchall()
    minsource_name = list(myresult1)
    minsource_namee = minsource_name[0][0]

    if minsource_namee =='':
        minsource_namee='BLANK'

    mycursor.execute('SELECT MAX(source_name) FROM available_trips')  # Min id value
    myresult1 = mycursor.fetchall()
    maxsource_name = list(myresult1)
    maxsource_namee = maxsource_name[0][0]

    sheet['D10'] = minsource_namee
    sheet['E10'] = maxsource_namee



    # attribute4:destination


    mycursor.execute('SELECT COUNT(DISTINCT destination) FROM available_trips')
    myresult4 = mycursor.fetchall()
    destinationcount = list(myresult4)
    destinationcountt = destinationcount[0][0]

    # globalistcount[2] = destinationcountt

    mycursor.execute('SELECT MIN(destination) FROM available_trips')
    myresult4 = mycursor.fetchall()
    mindestination = list(myresult4)
    mindestinationn = mindestination[0][0]

    if mindestinationn =='':
        mindestinationn='BLANK'

    mycursor.execute('SELECT MAX(destination) FROM available_trips')
    myresult4 = mycursor.fetchall()
    maxdestination = list(myresult4)
    maxdestinationn = maxdestination[0][0]

    sheet['D11'] = mindestinationn
    sheet['E11'] = maxdestinationn



    # attribute5:destination_name


    mycursor.execute('SELECT COUNT(DISTINCT destination_name) FROM available_trips')
    myresult5 = mycursor.fetchall()
    destinationnamecount = list(myresult5)
    destinationnamecountt = destinationnamecount[0][0]

    mycursor.execute('SELECT MIN(destination_name) FROM available_trips')
    myresult4 = mycursor.fetchall()
    mindestinationname = list(myresult4)
    mindestinationnamee = mindestinationname[0][0]

    if mindestinationnamee == '':
        mindestinationnamee = 'BLANK'

    mycursor.execute('SELECT MAX(destination_name) FROM available_trips')
    myresult4 = mycursor.fetchall()
    maxdestinationname = list(myresult4)
    maxdestinationnamee = maxdestinationname[0][0]

    sheet['D12'] = mindestinationnamee
    sheet['E12'] = maxdestinationnamee




    # attribute6:travels
    mycursor.execute('SELECT COUNT(DISTINCT travels) FROM available_trips')
    myresult6 = mycursor.fetchall()
    travelscount = list(myresult6)
    travelscountt = travelscount[0][0]

    mycursor.execute('SELECT MIN(travels) FROM available_trips')
    myresult4 = mycursor.fetchall()
    mintravels = list(myresult4)
    mintravelss = mintravels[0][0]

    if mintravelss == '':
        mintravelss = 'BLANK'

    mycursor.execute('SELECT MAX(travels) FROM available_trips')
    myresult4 = mycursor.fetchall()
    maxtravels = list(myresult4)
    maxtravelss = maxtravels[0][0]

    sheet['D13'] = mintravelss
    sheet['E13'] = maxtravelss



    # attrinute7 AC



    mycursor.execute('SELECT COUNT(DISTINCT AC) FROM available_trips')
    myresult7 = mycursor.fetchall()
    ACcount = list(myresult7)
    ACcountt = ACcount[0][0]

    mycursor.execute('SELECT MIN(AC) FROM available_trips')
    myresult7 = mycursor.fetchall()
    minac = list(myresult7)
    minacc = minac[0][0]

    if minacc == '':
        minacc = 'BLANK'

    mycursor.execute('SELECT MAX(AC) FROM available_trips')
    myresult7 = mycursor.fetchall()
    maxac = list(myresult7)
    maxacc = maxac[0][0]

    sheet['D14'] = minacc
    sheet['E14'] = maxacc



    # attribute7:arrivaltime



    mycursor.execute('SELECT COUNT(DISTINCT arrivalTime) FROM available_trips')
    myresult8 = mycursor.fetchall()
    arrivalTimecount = list(myresult8)
    arrivalTimecountt = arrivalTimecount[0][0]

    mycursor.execute('SELECT MIN(arrivalTime) FROM available_trips')
    myresult8 = mycursor.fetchall()
    minarrivalTime = list(myresult8)
    minarrivalTimee = minarrivalTime[0][0]

    if minarrivalTimee =='':
        minarrivalTimee="BLANK"

    mycursor.execute('SELECT MAX(arrivalTime) FROM available_trips')
    myresult8 = mycursor.fetchall()
    maxarrivalTime = list(myresult8)
    maxarrivalTimee = maxarrivalTime[0][0]

    sheet['D15'] = minarrivalTimee
    sheet['E15'] = maxarrivalTimee







    # attribute9:availCatCard
    mycursor.execute('SELECT COUNT(DISTINCT availCatCard) FROM available_trips')
    myresult9 = mycursor.fetchall()
    availCatCardcount = list(myresult9)
    availCatCardcountt = availCatCardcount[0][0]

    mycursor.execute('SELECT MIN(availCatCard) FROM available_trips')
    myresult9 = mycursor.fetchall()
    minavailCatCard = list(myresult9)
    minavailCatCardd = minavailCatCard[0][0]

    if minavailCatCardd == '':
        minavailCatCardd = 'BLANK'

    mycursor.execute('SELECT MAX(availCatCard) FROM available_trips')
    myresult9 = mycursor.fetchall()
    maxavailCatCard = list(myresult9)
    maxavailCatCardd = maxavailCatCard[0][0]

    sheet['D16'] = minavailCatCardd
    sheet['E16'] = maxavailCatCardd







    # attribute10:availSrCitizen
    mycursor.execute('SELECT COUNT(DISTINCT availSrCitizen) FROM available_trips')
    myresult10 = mycursor.fetchall()
    availSrCitizencount = list(myresult10)
    availSrCitizencountt = availSrCitizencount[0][0]

    mycursor.execute('SELECT MIN(availSrCitizen) FROM available_trips')
    myresult10 = mycursor.fetchall()
    minavailSrCitizen = list(myresult10)
    minavailSrCitizenn = minavailSrCitizen[0][0]

    if minavailSrCitizenn == '':
        minavailSrCitizenn = 'BLANK'

    mycursor.execute('SELECT MAX(availSrCitizen) FROM available_trips')
    myresult10 = mycursor.fetchall()
    maxavailSrCitizen = list(myresult10)
    maxavailSrCitizenn = maxavailSrCitizen[0][0]

    sheet['D17'] = minavailSrCitizenn
    sheet['E17'] = maxavailSrCitizenn




    # attribute11:availableSeats
    mycursor.execute('SELECT COUNT(DISTINCT availableSeats) FROM available_trips')
    myresult11 = mycursor.fetchall()
    availableSeatscount = list(myresult11)
    availableSeatscountt = availableSeatscount[0][0]

    mycursor.execute('SELECT MIN(availableSeats) FROM available_trips')
    myresult11 = mycursor.fetchall()
    minavailableSeats = list(myresult11)
    minavailableSeatss = minavailableSeats[0][0]

    if minavailableSeatss == '':
        minavailableSeatss = 'BLANK'

    mycursor.execute('SELECT MAX(availableSeats) FROM available_trips')
    myresult11 = mycursor.fetchall()
    maxavailableSeats = list(myresult11)
    maxavailableSeatss = maxavailableSeats[0][0]

    sheet['D18'] = minavailableSeatss
    sheet['E18'] = maxavailableSeatss



    # attribute12 :avlWindowSeats
    mycursor.execute('SELECT COUNT(DISTINCT avlWindowSeats) FROM available_trips')
    myresult12 = mycursor.fetchall()
    avlWindowSeatscount = list(myresult12)
    avlWindowSeatscountt = avlWindowSeatscount[0][0]

    mycursor.execute('SELECT MIN(avlWindowSeats) FROM available_trips')
    myresult12 = mycursor.fetchall()
    minavlWindowSeats = list(myresult12)
    minavlWindowSeatss = minavlWindowSeats[0][0]

    if minavlWindowSeatss == '':
        minavlWindowSeatss = 'BLANK'

    mycursor.execute('SELECT MAX(avlWindowSeats) FROM available_trips')
    myresult12 = mycursor.fetchall()
    maxavlWindowSeats = list(myresult12)
    maxavlWindowSeatss = maxavlWindowSeats[0][0]

    sheet['D19'] = minavlWindowSeatss
    sheet['E19'] = maxavlWindowSeatss



    # attribute13:	boardingTimes
    mycursor.execute('SELECT COUNT(DISTINCT boardingTimes) FROM available_trips')
    myresult13 = mycursor.fetchall()
    boardingTimescount = list(myresult13)
    boardingTimescountt = boardingTimescount[0][0]

    mycursor.execute('SELECT MIN(boardingTimes) FROM available_trips')
    myresult12 = mycursor.fetchall()
    minaboardingTimes= list(myresult12)
    minaboardingTimess = minaboardingTimes[0][0]


    if minaboardingTimess == '':
        minaboardingTimess = 'BLANK'

    mycursor.execute('SELECT MAX(boardingTimes) FROM available_trips')
    myresult12 = mycursor.fetchall()
    maxboardingTimes = list(myresult12)
    maxboardingTimess = maxboardingTimes[0][0]

    sheet['D20'] = minaboardingTimess
    sheet['E20'] = maxboardingTimess



    # attribute14:bookable
    mycursor.execute('SELECT COUNT(DISTINCT bookable) FROM available_trips')
    myresult14 = mycursor.fetchall()
    bookablecount = list(myresult14)
    bookablecountt = bookablecount[0][0]

    mycursor.execute('SELECT MIN(bookable) FROM available_trips')
    myresult14 = mycursor.fetchall()
    minbookable = list(myresult14)
    minbookablee = minbookable[0][0]

    if minbookablee == '':
        minbookablee = 'BLANK'

    mycursor.execute('SELECT MAX(bookable) FROM available_trips')
    myresult14 = mycursor.fetchall()
    maxbookable = list(myresult14)
    maxbookablee = maxbookable[0][0]

    sheet['D21'] = minbookablee
    sheet['E21'] = maxbookablee



    # attribute15:bpDpSeatLayout
    mycursor.execute('SELECT COUNT(DISTINCT bpDpSeatLayout) FROM available_trips')
    myresult15 = mycursor.fetchall()
    bpDpSeatLayoutcount = list(myresult15)
    bpDpSeatLayoutcountt = bpDpSeatLayoutcount[0][0]

    mycursor.execute('SELECT MIN(bpDpSeatLayout) FROM available_trips')
    myresult15 = mycursor.fetchall()
    minbpDpSeatLayout = list(myresult15)
    minbpDpSeatLayoutt = minbpDpSeatLayout[0][0]

    if minbpDpSeatLayoutt == '':
        minbpDpSeatLayoutt = 'BLANK'

    mycursor.execute('SELECT MAX(bpDpSeatLayout) FROM available_trips')
    myresult15 = mycursor.fetchall()
    maxbpDpSeatLayout = list(myresult15)
    maxbpDpSeatLayoutt = maxbpDpSeatLayout[0][0]

    sheet['D22'] = minbpDpSeatLayoutt
    sheet['E22'] = maxbpDpSeatLayoutt



    # attribute16:busImageCount
    mycursor.execute('SELECT COUNT(DISTINCT busImageCount) FROM available_trips')
    myresult16 = mycursor.fetchall()
    busImageCountcount = list(myresult16)
    busImageCountcountt = busImageCountcount[0][0]

    mycursor.execute('SELECT MIN(busImageCount) FROM available_trips')
    myresult16 = mycursor.fetchall()
    minbusImageCount = list(myresult16)
    minbusImageCountt = minbusImageCount[0][0]

    if minbusImageCountt == '':
        minbusImageCountt = 'BLANK'

    mycursor.execute('SELECT MAX(busImageCount) FROM available_trips')
    myresult16 = mycursor.fetchall()
    maxbusImageCount = list(myresult16)
    maxbusImageCountt = maxbusImageCount[0][0]

    sheet['D23'] = minbusImageCountt
    sheet['E23'] = maxbusImageCountt


    # attribute17:busServiceId
    mycursor.execute('SELECT COUNT(DISTINCT  busServiceId) FROM available_trips')
    myresult17 = mycursor.fetchall()
    busServiceIdcount = list(myresult17)
    busServiceIdcountt = busServiceIdcount[0][0]

    mycursor.execute('SELECT MIN(busServiceId) FROM available_trips')
    myresult17 = mycursor.fetchall()
    minbusServiceId = list(myresult17)
    minbusServiceIdd = minbusServiceId[0][0]

    if minbusServiceIdd=='':
        minbusServiceIdd= "BLANK"


    mycursor.execute('SELECT MAX(busServiceId) FROM available_trips')
    myresult17 = mycursor.fetchall()
    maxbusServiceId = list(myresult17)
    maxbusServiceIdd = maxbusServiceId[0][0]

    sheet['D24'] = minbusServiceIdd
    sheet['E24'] = maxbusServiceIdd




        # attribute18:busType
    mycursor.execute('SELECT COUNT(DISTINCT busType) FROM available_trips')
    myresult18 = mycursor.fetchall()
    busTypecount = list(myresult18)
    busTypecountt = busTypecount[0][0]

    mycursor.execute('SELECT MIN(busType) FROM available_trips')
    myresult17 = mycursor.fetchall()
    minbusType = list(myresult17)
    minbusTypee = minbusType[0][0]

    if minbusTypee == '':
        minbusTypee = 'BLANK'

    mycursor.execute('SELECT MAX(busType) FROM available_trips')
    myresult17 = mycursor.fetchall()
    maxbusType = list(myresult17)
    maxbusTypee = maxbusType[0][0]

    sheet['D25'] = minbusTypee
    sheet['E25'] = maxbusTypee




    # attribute19:busTypeId
    mycursor.execute('SELECT COUNT(DISTINCT busTypeId) FROM available_trips')
    myresult19 = mycursor.fetchall()
    busTypeIdcount = list(myresult19)
    busTypeIdcountt = busTypeIdcount[0][0]

    mycursor.execute('SELECT MIN(busTypeId) FROM available_trips')
    myresult17 = mycursor.fetchall()
    minbusTypeId = list(myresult17)
    minbusTypeIdd = minbusTypeId[0][0]

    if minbusTypeIdd == '':
        minbusTypeIdd = 'BLANK'

    mycursor.execute('SELECT MAX(busTypeId) FROM available_trips')
    myresult17 = mycursor.fetchall()
    maxbusTypeId = list(myresult17)
    maxbusTypeId = maxbusTypeId[0][0]

    sheet['D26'] = minbusTypeIdd
    sheet['E26'] = maxbusTypeId




    # attribute20:cancellationPolicy
    mycursor.execute('SELECT COUNT(DISTINCT cancellationPolicy) FROM available_trips')
    myresult20 = mycursor.fetchall()
    cancellationPolicycount = list(myresult20)
    cancellationPolicycountt = cancellationPolicycount[0][0]

    mycursor.execute('SELECT MIN(cancellationPolicy) FROM available_trips')
    myresult17 = mycursor.fetchall()
    mincancellationPolicy = list(myresult17)
    mincancellationPolicyy = mincancellationPolicy[0][0]

    if mincancellationPolicyy == '':
        mincancellationPolicyy = 'BLANK'

    mycursor.execute('SELECT MAX(cancellationPolicy) FROM available_trips')
    myresult17 = mycursor.fetchall()
    maxcancellationPolicy = list(myresult17)
    maxcancellationPolicyy = maxcancellationPolicy[0][0]

    sheet['D27'] = mincancellationPolicyy
    sheet['E27'] = maxcancellationPolicyy




    # attribute21:departureTime
    mycursor.execute('SELECT COUNT(DISTINCT departureTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    departureTimecount = list(myresult21)
    departureTimecountt = departureTimecount[0][0]

    mycursor.execute('SELECT MIN(departureTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    mindepartureTime = list(myresult21)
    mindepartureTimee = mindepartureTime[0][0]


    if mindepartureTimee == '':
        mindepartureTimee = 'BLANK'

    mycursor.execute('SELECT MAX(departureTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxdepartureTime = list(myresult21)
    maxdepartureTime = maxdepartureTime[0][0]

    sheet['D28'] = mindepartureTimee
    sheet['E28'] = maxdepartureTime



    # attribute22:doj
    mycursor.execute('SELECT COUNT(DISTINCT doj) FROM available_trips')
    myresult22 = mycursor.fetchall()
    dojcount = list(myresult22)
    dojcountt = dojcount[0][0]

    mycursor.execute('select distinct doj from available_trips ORDER BY doj ASC')
    myresult22 = mycursor.fetchall()  # orderby min logic
    mindoj = [item for x in zip_longest(*myresult22) for item in x if item != -55 and item != '']
    mindojj = mindoj[0]

    if mindojj =='':
        mindojj = 'BLANK'

    mycursor.execute('select distinct doj from available_trips ORDER BY doj DESC')
    myresult22 = mycursor.fetchall()  ##orderby man logic
    maxdoj = [item for x in zip_longest(*myresult22) for item in x if item != -55 and item != '']
    maxdojj = maxdoj[0]

    sheet['D29'] = mindojj
    sheet['E29'] = maxdojj



    # attribute23:dropPointMandatory
    mycursor.execute('SELECT COUNT(DISTINCT dropPointMandatory) FROM available_trips')
    myresult23 = mycursor.fetchall()
    dropPointMandatorycount = list(myresult23)
    dropPointMandatorycountt = dropPointMandatorycount[0][0]

    mycursor.execute('SELECT MIN(dropPointMandatory) FROM available_trips')
    myresult21 = mycursor.fetchall()
    mindropPointMandatory = list(myresult21)
    mindropPointMandatoryy = mindropPointMandatory[0][0]

    if mindropPointMandatoryy == '':
        mindropPointMandatoryy = 'BLANK'

    mycursor.execute('SELECT MAX(dropPointMandatory) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxdropPointMandatory = list(myresult21)
    maxdropPointMandatoryy = maxdropPointMandatory[0][0]

    sheet['D30'] = mindropPointMandatoryy
    sheet['E30'] = maxdropPointMandatoryy




    # attribute24:	droppingTimes
    mycursor.execute('SELECT COUNT(DISTINCT droppingTimes) FROM available_trips')
    myresult24 = mycursor.fetchall()
    droppingTimescount = list(myresult24)
    droppingTimescountt = droppingTimescount[0][0]

    mycursor.execute('SELECT MIN(droppingTimes) FROM available_trips')
    myresult21 = mycursor.fetchall()
    mindroppingTimes = list(myresult21)
    mindroppingTimess = mindroppingTimes[0][0]

    if mindroppingTimes=='':
        mindroppingTimes='BLANK'


    mycursor.execute('SELECT MAX(droppingTimes) FROM available_trips')
    myresult21 = mycursor.fetchall()
    mindroppingTimes = list(myresult21)
    mindroppingTimess = mindroppingTimes[0][0]

    sheet['D31'] = mindroppingTimess
    sheet['E31'] = mindroppingTimess



    # attribute25:		fareDetails
    mycursor.execute('SELECT COUNT(DISTINCT fareDetails) FROM available_trips')
    myresult25 = mycursor.fetchall()
    fareDetailscount = list(myresult25)
    fareDetailscountt = fareDetailscount[0][0]

    mycursor.execute('select distinct fareDetails from available_trips ORDER BY fareDetails ASC')
    myresult25 = mycursor.fetchall()  # orderby min logic
    minfareDetails = [item for x in zip_longest(*myresult25) for item in x if item != -55 and item != '']
    minfareDetailss = minfareDetails[0]


    if minfareDetailss == '':
        minfareDetailss = 'BLANK'

    mycursor.execute('select distinct fareDetails from available_trips ORDER BY fareDetails DESC')
    myresult25 = mycursor.fetchall()  # orderby min logic
    maxfareDetails = [item for x in zip_longest(*myresult25) for item in x if item != -55 and item != '']
    maxfareDetailss = maxfareDetails[0]

    sheet['D32'] = minfareDetailss
    sheet['E32'] = maxfareDetailss



    # attribute26:	fares      #min max logic remain
    mycursor.execute('SELECT COUNT(DISTINCT fares) FROM available_trips')
    myresult26 = mycursor.fetchall()
    farescount = list(myresult26)
    farescountt = farescount[0][0]

    mycursor.execute('SELECT MIN(fares) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minfares = list(myresult21)
    minfaress = minfares[0][0]

    if minfaress =='':
        minfaress = 'BLANK'



    mycursor.execute('SELECT MAX(fares) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxfares = list(myresult21)
    maxfaress = maxfares[0][0]

    sheet['D33'] = minfaress
    sheet['E33'] = maxfaress

    # attribute27:	flatComApplicable
    mycursor.execute('SELECT COUNT(DISTINCT flatComApplicable) FROM available_trips')
    myresult27 = mycursor.fetchall()
    flatComApplicablecount = list(myresult27)
    flatComApplicablecountt = flatComApplicablecount[0][0]

    mycursor.execute('SELECT MIN(flatComApplicable) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minflatComApplicable = list(myresult21)
    minflatComApplicablee = minflatComApplicable[0][0]


    if minflatComApplicablee == '':
        minflatComApplicablee = 'BLANK'

    mycursor.execute('SELECT MAX(flatComApplicable) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxflatComApplicable = list(myresult21)
    maxflatComApplicablee = maxflatComApplicable[0][0]

    sheet['D34'] = minflatComApplicablee
    sheet['E34'] = maxflatComApplicablee



    # attribute28:	gdsCommission
    mycursor.execute('SELECT COUNT(DISTINCT gdsCommission) FROM available_trips')
    myresult28 = mycursor.fetchall()
    gdsCommissioncount = list(myresult28)
    gdsCommissioncountt = gdsCommissioncount[0][0]

    mycursor.execute('SELECT MIN(gdsCommission) FROM available_trips')
    myresult21 = mycursor.fetchall()
    mingdsCommission = list(myresult21)
    mingdsCommissionn = mingdsCommission[0][0]

    if mingdsCommissionn == '':
        mingdsCommissionn = 'BLANK'

    mycursor.execute('SELECT MAX(gdsCommission) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxgdsCommission = list(myresult21)
    maxgdsCommissionn = maxgdsCommission[0][0]



    sheet['D35'] = mingdsCommissionn
    sheet['E35'] = maxgdsCommissionn

    # attribute29:		idProofRequired
    mycursor.execute('SELECT COUNT(DISTINCT idProofRequired) FROM available_trips')
    myresult29 = mycursor.fetchall()
    idProofRequiredcount = list(myresult29)
    idProofRequiredcountt = idProofRequiredcount[0][0]

    mycursor.execute('SELECT MIN(idProofRequired) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minidProofRequired = list(myresult21)
    minidProofRequiredd = minidProofRequired[0][0]

    if minidProofRequiredd == '':
        minidProofRequiredd = 'BLANK'

    mycursor.execute('SELECT MAX(idProofRequired) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxidProofRequired = list(myresult21)
    maxidProofRequiredd = maxidProofRequired[0][0]

    sheet['D36'] = minidProofRequiredd
    sheet['E36'] = maxidProofRequiredd


    # attribute30:		liveTrackingAvailable
    mycursor.execute('SELECT COUNT(DISTINCT liveTrackingAvailable) FROM available_trips')
    myresult30 = mycursor.fetchall()
    liveTrackingAvailablecount = list(myresult30)
    liveTrackingAvailablecountt = liveTrackingAvailablecount[0][0]

    mycursor.execute('SELECT MIN(liveTrackingAvailable) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minliveTrackingAvailable = list(myresult21)
    minliveTrackingAvailablee = minliveTrackingAvailable[0][0]

    if minliveTrackingAvailablee == '':
        minliveTrackingAvailablee = 'BLANK'

    mycursor.execute('SELECT MAX(liveTrackingAvailable) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxliveTrackingAvailable = list(myresult21)
    maxliveTrackingAvailablee = maxliveTrackingAvailable[0][0]

    sheet['D37'] = minliveTrackingAvailablee
    sheet['E37'] = maxliveTrackingAvailablee


    # attribute31:		maxSeatsPerTicket
    mycursor.execute('SELECT COUNT(DISTINCT maxSeatsPerTicket) FROM available_trips')
    myresult31 = mycursor.fetchall()
    maxSeatsPerTicketcount = list(myresult31)
    maxSeatsPerTicketcountt = maxSeatsPerTicketcount[0][0]

    mycursor.execute('SELECT MIN(maxSeatsPerTicket) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minmaxSeatsPerTicket = list(myresult21)
    minmaxSeatsPerTickett = minmaxSeatsPerTicket[0][0]

    if minmaxSeatsPerTickett == '':
        minmaxSeatsPerTickett = 'BLANK'

    mycursor.execute('SELECT MAX(maxSeatsPerTicket) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxmaxSeatsPerTicket = list(myresult21)
    maxmaxSeatsPerTickett = maxmaxSeatsPerTicket[0][0]

    sheet['D38'] = minmaxSeatsPerTickett
    sheet['E38'] = maxmaxSeatsPerTickett




    # attribute32:	nonAC
    mycursor.execute('SELECT COUNT(DISTINCT nonAC) FROM available_trips')
    myresult32 = mycursor.fetchall()
    nonACcount = list(myresult32)
    nonACcountt = nonACcount[0][0]

    mycursor.execute('SELECT MIN(nonAC) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minnonAC = list(myresult21)
    minnonACC = minnonAC[0][0]

    if minnonACC == '':
        minnonACC = 'BLANK'

    mycursor.execute('SELECT MAX(nonAC) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxnonAC = list(myresult21)
    maxnonACC = maxnonAC[0][0]

    sheet['D39'] = minnonACC
    sheet['E39'] = maxnonACC

    # attribute33:	operator
    mycursor.execute('SELECT COUNT(DISTINCT operator) FROM available_trips')
    myresult33 = mycursor.fetchall()
    operatorcount = list(myresult33)
    operatorcountt = operatorcount[0][0]

    mycursor.execute('SELECT MIN(operator) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minoperator = list(myresult21)
    minoperatorr = minoperator[0][0]

    if minoperatorr == '':
        minoperatorr = 'BLANK'

    mycursor.execute('SELECT MAX(operator) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxoperator = list(myresult21)
    maxoperatorr = maxoperator[0][0]

    sheet['D40'] = minoperatorr
    sheet['E40'] = maxoperatorr




    # attribute34:	otgEnabled
    mycursor.execute('SELECT COUNT(DISTINCT otgEnabled) FROM available_trips')
    myresult34 = mycursor.fetchall()
    otgEnabledcount = list(myresult34)
    otgEnabledcountt = otgEnabledcount[0][0]

    mycursor.execute('SELECT MIN(otgEnabled) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minotgEnabled = list(myresult21)
    minotgEnabledd = minotgEnabled[0][0]

    if minotgEnabledd == '':
        minotgEnabledd = 'BLANK'

    mycursor.execute('SELECT MAX(otgEnabled) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxotgEnabled = list(myresult21)
    maxotgEnabledd = maxotgEnabled[0][0]

    sheet['D41'] = minotgEnabledd
    sheet['E41'] = maxotgEnabledd

    # attribute35:	otgPolicy
    mycursor.execute('SELECT COUNT(DISTINCT otgPolicy) FROM available_trips')
    myresult35 = mycursor.fetchall()
    otgPolicycount = list(myresult35)
    otgPolicycountt = otgPolicycount[0][0]

    mycursor.execute('SELECT MIN(otgPolicy) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minotgPolicy = list(myresult21)
    minotgPolicyy = minotgPolicy[0][0]

    if minotgPolicyy == '':
        minotgPolicyy = 'BLANK'

    mycursor.execute('SELECT MAX(otgPolicy) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxotgPolicy = list(myresult21)
    maxotgPolicyy = maxotgPolicy[0][0]

    sheet['D42'] = minotgPolicyy
    sheet['E42'] = maxotgPolicyy

    # attribute36:	partialCancellationAllowed
    mycursor.execute('SELECT COUNT(DISTINCT partialCancellationAllowed) FROM available_trips')
    myresult36 = mycursor.fetchall()
    partialCancellationAllowedcount = list(myresult36)
    partialCancellationAllowedcountt = partialCancellationAllowedcount[0][0]

    mycursor.execute('SELECT MIN(partialCancellationAllowed) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minpartialCancellationAllowed = list(myresult21)
    minpartialCancellationAllowedd = minpartialCancellationAllowed[0][0]

    if minpartialCancellationAllowedd == '':
        minpartialCancellationAllowedd = 'BLANK'

    mycursor.execute('SELECT MAX(partialCancellationAllowed) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxpartialCancellationAllowed = list(myresult21)
    maxpartialCancellationAllowedd = maxpartialCancellationAllowed[0][0]

    sheet['D43'] = minpartialCancellationAllowedd
    sheet['E43'] = maxpartialCancellationAllowedd




    # attribute37:	partnerBaseCommission
    mycursor.execute('SELECT COUNT(DISTINCT partnerBaseCommission) FROM available_trips')
    myresult37 = mycursor.fetchall()
    partnerBaseCommissioncount = list(myresult37)
    partnerBaseCommissioncountt = partnerBaseCommissioncount[0][0]

    mycursor.execute('SELECT MIN(partnerBaseCommission) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minpartnerBaseCommission= list(myresult21)
    minpartnerBaseCommissionn = minpartnerBaseCommission[0][0]

    if minpartnerBaseCommissionn == '':
        minpartnerBaseCommissionn = 'BLANK'

    mycursor.execute('SELECT MAX(partnerBaseCommission) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxpartnerBaseCommission= list(myresult21)
    maxpartnerBaseCommissionn = maxpartnerBaseCommission[0][0]

    sheet['D44'] = minpartnerBaseCommissionn
    sheet['E44'] = maxpartnerBaseCommissionn


    # attribute38:	primaryPaxCancellable
    mycursor.execute('SELECT COUNT(DISTINCT primaryPaxCancellable) FROM available_trips')
    myresult38 = mycursor.fetchall()
    primaryPaxCancellablecount = list(myresult38)
    primaryPaxCancellablecountt = primaryPaxCancellablecount[0][0]


    mycursor.execute('SELECT MIN(primaryPaxCancellable) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minprimaryPaxCancellable= list(myresult21)
    minprimaryPaxCancellablee = minprimaryPaxCancellable[0][0]

    if minprimaryPaxCancellablee == '':
        minprimaryPaxCancellablee = 'BLANK'

    mycursor.execute('SELECT MAX(primaryPaxCancellable) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxprimaryPaxCancellable= list(myresult21)
    maxprimaryPaxCancellablee = maxprimaryPaxCancellable[0][0]

    sheet['D45'] = minprimaryPaxCancellablee
    sheet['E45'] = maxprimaryPaxCancellablee



    # attribute39:	routeId
    mycursor.execute('SELECT COUNT(DISTINCT routeId) FROM available_trips')
    myresult39 = mycursor.fetchall()
    routeIdcount = list(myresult39)
    routeIdcountt = routeIdcount[0][0]

    mycursor.execute('SELECT MIN(routeId) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minrouteId = list(myresult21)
    minrouteIdd = minrouteId[0][0]

    if minrouteIdd == '':
        minrouteIdd = 'BLANK'

    mycursor.execute('SELECT MAX(routeId) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxrouteId = list(myresult21)
    maxrouteIdd = maxrouteId[0][0]

    sheet['D46'] = minrouteIdd
    sheet['E46'] = maxrouteIdd


    # attribute40:	rtc
    mycursor.execute('SELECT COUNT(DISTINCT rtc) FROM available_trips')
    myresult40 = mycursor.fetchall()
    rtccount = list(myresult40)
    rtccountt = rtccount[0][0]

    mycursor.execute('SELECT MIN(rtc) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minrtc = list(myresult21)
    minrtcc = minrtc[0][0]

    if minrtcc == '':
        minrtcc = 'BLANK'

    mycursor.execute('SELECT MAX(rtc) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxrtc= list(myresult21)
    maxrtcc = maxrtc[0][0]

    sheet['D47'] = minrtcc
    sheet['E47'] = maxrtcc

    # attribute41:		seater
    mycursor.execute('SELECT COUNT(DISTINCT seater) FROM available_trips')
    myresult41 = mycursor.fetchall()
    seatercount = list(myresult41)
    seatercountt = seatercount[0][0]

    mycursor.execute('SELECT MIN(seater) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minseater = list(myresult21)
    minseaterr = minseater[0][0]

    mycursor.execute('SELECT MAX(seater) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxseater = list(myresult21)
    maxseaterr = maxseater[0][0]

    sheet['D48'] = minseaterr
    sheet['E48'] = maxseaterr


    # attribute42:	selfInventory
    mycursor.execute('SELECT COUNT(DISTINCT selfInventory) FROM available_trips')
    myresult42 = mycursor.fetchall()
    selfInventorycount = list(myresult42)
    selfInventorycountt = selfInventorycount[0][0]

    mycursor.execute('SELECT MIN(selfInventory) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minselfInventory = list(myresult21)
    minselfInventoryy = minselfInventory[0][0]

    if minselfInventoryy == '':
        minselfInventoryy = 'BLANK'

    mycursor.execute('SELECT MAX(selfInventory) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxselfInventory = list(myresult21)
    maxselfInventoryy = maxselfInventory[0][0]

    sheet['D49'] = minselfInventoryy
    sheet['E49'] = maxselfInventoryy




    # attribute43:	singleLadies
    mycursor.execute('SELECT COUNT(DISTINCT singleLadies) FROM available_trips')
    myresult43 = mycursor.fetchall()
    singleLadiescount = list(myresult43)
    singleLadiescountt = singleLadiescount[0][0]

    mycursor.execute('SELECT MIN(singleLadies) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minsingleLadies = list(myresult21)
    minsingleLadiess = minsingleLadies[0][0]

    if minsingleLadiess == '':
        minsingleLadiess = 'BLANK'

    mycursor.execute('SELECT MAX(singleLadies) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxsingleLadies = list(myresult21)
    maxsingleLadiess = maxsingleLadies[0][0]

    sheet['D50'] = minsingleLadiess
    sheet['E50'] = maxsingleLadiess








    # attribute44:	sleeper
    mycursor.execute('SELECT COUNT(DISTINCT sleeper) FROM available_trips')
    myresult44 = mycursor.fetchall()
    sleepercount = list(myresult44)
    sleepercountt = sleepercount[0][0]

    mycursor.execute('SELECT MIN(sleeper) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minsleeper= list(myresult21)
    minsleeperr = minsleeper[0][0]

    if minsleeperr == '':
        minsleeperr = 'BLANK'

    mycursor.execute('SELECT MAX(sleeper) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxsleeper = list(myresult21)
    maxsleeperr = maxsleeper[0][0]

    sheet['D51'] = minsleeperr
    sheet['E51'] = maxsleeperr




    # attribute45:	tatkalTime
    mycursor.execute('SELECT COUNT(DISTINCT tatkalTime) FROM available_trips')
    myresult45 = mycursor.fetchall()
    tatkalTimecount = list(myresult45)
    tatkalTimecountt = tatkalTimecount[0][0]

    mycursor.execute('SELECT MIN(tatkalTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    mintatkalTime = list(myresult21)
    mintatkalTimee = mintatkalTime[0][0]


    if mintatkalTimee == '':
        mintatkalTimee = 'BLANK'

    mycursor.execute('SELECT MAX(tatkalTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxtatkalTime = list(myresult21)
    maxtatkalTimee = maxtatkalTime[0][0]

    sheet['D52'] = mintatkalTimee
    sheet['E52'] = maxtatkalTimee




    # attribute46:	vehicleType
    mycursor.execute('SELECT COUNT(DISTINCT vehicleType) FROM available_trips')
    myresult46 = mycursor.fetchall()
    vehicleTypecount = list(myresult46)
    vehicleTypecountt = vehicleTypecount[0][0]

    mycursor.execute('SELECT MIN(vehicleType) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minvehicleType = list(myresult21)
    minvehicleTypee = minvehicleType[0][0]

    if minvehicleTypee == '':
        minvehicleTypee = 'BLANK'

    mycursor.execute('SELECT MAX(vehicleType) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxvehicleType = list(myresult21)
    maxvehicleTypee = maxvehicleType[0][0]

    sheet['D53'] = minvehicleTypee
    sheet['E53'] = maxvehicleTypee

    # attribute47:		viaRoutes
    mycursor.execute('SELECT COUNT(DISTINCT viaRoutes) FROM available_trips')
    myresult47 = mycursor.fetchall()
    viaRoutescount = list(myresult47)
    viaRoutescountt = viaRoutescount[0][0]

    mycursor.execute('SELECT MIN(viaRoutes) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minviaRoutes = list(myresult21)
    minviaRoutess = minviaRoutes[0][0]

    if minviaRoutess == '':
        minviaRoutess = 'BLANK'

    mycursor.execute('SELECT MAX(viaRoutes) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxviaRoutes = list(myresult21)
    maxviaRoutess = maxviaRoutes[0][0]

    sheet['D54'] = minviaRoutess
    sheet['E54'] = maxviaRoutess



    # attribute48:	zeroCancellationTime
    mycursor.execute('SELECT COUNT(DISTINCT zeroCancellationTime) FROM available_trips')
    myresult48 = mycursor.fetchall()
    zeroCancellationTimecount = list(myresult48)
    zeroCancellationTimecountt = zeroCancellationTimecount[0][0]

    mycursor.execute('SELECT MIN(zeroCancellationTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minzeroCancellationTime = list(myresult21)
    minzeroCancellationTimee = minzeroCancellationTime[0][0]

    if minzeroCancellationTimee == '':
        minzeroCancellationTimee = 'BLANK'

    mycursor.execute('SELECT MAX(zeroCancellationTime) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxminzeroCancellationTime = list(myresult21)
    maxminzeroCancellationTimee = maxminzeroCancellationTime[0][0]

    sheet['D55'] = minzeroCancellationTimee
    sheet['E55'] = maxminzeroCancellationTimee



    # attribute49:		mTicketEnabled
    mycursor.execute('SELECT COUNT(DISTINCT mTicketEnabled) FROM available_trips')
    myresult49 = mycursor.fetchall()
    mTicketEnabledcount = list(myresult49)
    mTicketEnabledcountt = mTicketEnabledcount[0][0]

    mycursor.execute('SELECT MIN(mTicketEnabled) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minmTicketEnabled = list(myresult21)
    minmTicketEnabled = minmTicketEnabled[0][0]


    if minmTicketEnabled == '':
        minmTicketEnabled = 'BLANK'

    mycursor.execute('SELECT MAX(mTicketEnabled) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxmTicketEnabled = list(myresult21)
    maxmTicketEnabledd = maxmTicketEnabled[0][0]

    sheet['D56'] = minmTicketEnabled
    sheet['E56'] = maxmTicketEnabledd

    # attribute50:		sd_id
    mycursor.execute('SELECT COUNT(DISTINCT sd_id) FROM available_trips')
    myresult49 = mycursor.fetchall()
    sd_idcount = list(myresult49)
    sd_idcountt = sd_idcount[0][0]

    mycursor.execute('SELECT MIN(sd_id) FROM available_trips')
    myresult21 = mycursor.fetchall()
    minsd_id = list(myresult21)
    minsd_idd = minsd_id[0][0]

    if minsd_id == '':
        minsd_id = 'BLANK'

    mycursor.execute('SELECT MAX(sd_id) FROM available_trips')
    myresult21 = mycursor.fetchall()
    maxsd_id = list(myresult21)
    maxsd_idd = maxsd_id[0][0]

    sheet['D57'] = minsd_idd
    sheet['E57'] = maxsd_idd


    # attribute51:          createDt
    mycursor.execute('SELECT COUNT(DISTINCT  createDt) FROM available_trips')
    myresult50 = mycursor.fetchall()
    createDtcount = list(myresult50)
    createDtcountt = createDtcount[0][0]

    mycursor.execute('select distinct createDt from available_trips ORDER BY createDt ASC')
    myresult50 = mycursor.fetchall()  # orderby min logic
    mincreateDt = [item for x in zip_longest(*myresult50) for item in x if item != -55 and item != '']
    mincreateDtt = mincreateDt[0]

    if mincreateDtt == '':
        mincreateDtt = 'BLANK'

    mycursor.execute('select distinct   createDt from available_trips ORDER BY  createDt DESC')
    myresult50 = mycursor.fetchall()  # orderby max logic
    maxcreateDt = [item for x in zip_longest(*myresult50) for item in x if item != -55 and item != '']
    maxcreateDtt = maxcreateDt[0]

    sheet['D58'] = mincreateDtt
    sheet['E58'] = maxcreateDtt



    # attribute52:		created_date
    mycursor.execute('SELECT COUNT(DISTINCT created_date) FROM available_trips')
    myresult50 = mycursor.fetchall()
    created_datecount = list(myresult50)
    created_datecountt = created_datecount[0][0]

    mycursor.execute('select distinct created_date from available_trips ORDER BY created_date ASC')
    myresult50 = mycursor.fetchall()  # orderby min logic
    mincreated_date = [item for x in zip_longest(*myresult50) for item in x if item != -55 and item != '']
    mincreated_datee = mincreated_date[0]

    if mincreated_datee == '':
        mincreated_datee = 'BLANK'

    mycursor.execute('select distinct created_date from available_trips ORDER BY created_date DESC')
    myresult50 = mycursor.fetchall()  # orderby max logic
    maxcreated_date = [item for x in zip_longest(*myresult50) for item in x if item != -55 and item != '']
    maxcreated_datee = maxcreated_date[0]

    sheet['D59'] = mincreated_datee
    sheet['E59'] = maxcreated_datee



    sheet['C6'] = "Min:"+ str(mincreated_datee)+","+"Max:" + str(maxcreated_datee)
    sheet['B1'] = "DB Name"
    sheet['B2'] = "TABLE NAME"
    sheet['B3'] = "TOTAL RECORDS"
    sheet['B4'] = "TOTAL BLANK RECORDS ACC. TO DOJ"
    sheet['B5'] = "DIFFRENCE(ACTUAL RECORDS) "
    sheet['B6'] = "CREATED DATE"
    sheet['A7'] = "Attribute.No"
    sheet['B7'] = "Attribute Name"
    sheet['D7'] = "Min Value"
    sheet['E7'] = "Max Value"
    sheet['F7'] = "DIFFERENCE VALUE"
    sheet['C7'] = "COUNT (DISTINCT)"


    # sheet['B3'] = sourcenamecountt
    sheet.cell(row=1, column=2).font = Font(size=15)
    sheet.cell(row=2, column=2).font = Font(size=15)
    sheet.cell(row=3, column=2).font = Font(size=15)
    sheet.cell(row=4, column=2).font = Font(size=15)
    sheet.cell(row=5, column=2).font = Font(size=15)
    sheet.cell(row=6, column=2).font = Font(size=15)

    sheet.cell(row=1, column=3).font = Font(size=12)
    sheet.cell(row=2, column=3).font = Font(size=12)
    sheet.cell(row=3, column=3).font = Font(size=12)
    sheet.cell(row=4, column=3).font = Font(size=12)
    sheet.cell(row=5, column=3).font = Font(size=12)
    sheet.cell(row=6, column=3).font = Font(size=12)
    sheet.cell(row=7, column=1).font = Font(size=15)
    sheet.cell(row=7, column=2).font = Font(size=15)
    sheet.cell(row=7, column=3).font = Font(size=15)
    sheet.cell(row=7, column=4).font = Font(size=15)
    sheet.cell(row=7, column=5).font = Font(size=15)


    # sheet['A1'].fill = PatternFill(bgColor="Orange", fill_type = "solid")
    # for x in range(32):
    #   sheet.cell(row=7, column=1+x).fill = PatternFill(bgColor="Orange", fill_type = "solid")


        # set the height of the row
    sheet.row_dimensions[7].height = 25
    # set the width of the column
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 45
    sheet.column_dimensions['C'].width = 55
    sheet.column_dimensions['D'].width = 45
    sheet.column_dimensions['E'].width = 45
    sheet.column_dimensions['F'].width = 30

    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 30
    sheet.column_dimensions['K'].width = 30
    sheet.column_dimensions['L'].width = 30
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 30
    sheet.column_dimensions['O'].width = 30
    sheet.column_dimensions['P'].width = 30
    sheet.column_dimensions['Q'].width = 30
    sheet.column_dimensions['R'].width = 30
    sheet.column_dimensions['S'].width = 30
    sheet.column_dimensions['T'].width = 30
    sheet.column_dimensions['U'].width = 30
    sheet.column_dimensions['V'].width = 30
    sheet.column_dimensions['W'].width = 30
    sheet.column_dimensions['X'].width = 30
    sheet.column_dimensions['Y'].width = 30
    sheet.column_dimensions['Z'].width = 30
    sheet.column_dimensions['AA'].width = 30
    sheet.column_dimensions['AB'].width = 30
    sheet.column_dimensions['AC'].width = 30
    sheet.column_dimensions['AD'].width = 30
    sheet.column_dimensions['AE'].width = 30
    sheet.column_dimensions['AF'].width = 30
    sheet.column_dimensions['AG'].width = 30
    sheet.column_dimensions['AH'].width = 30
    sheet.column_dimensions['AI'].width = 30
    sheet.column_dimensions['AJ'].width = 30

    numberlist = list(range(1, 53))
    for x in range(52):  # attribute number loop
        c3 = sheet.cell(row=x + 8, column=1)
        c3.value = numberlist[x]
    # sheet['F9'] = (driver.var9 - driver.var10)

    sheet['C8'] = idcountt
    sheet['C9'] = sourcecountt
    sheet['C10'] = sourcenamecountt
    sheet['C11'] = destinationcountt
    sheet['C12'] = destinationnamecountt
    sheet['C13'] = travelscountt
    sheet['C14'] = ACcountt
    sheet['C15'] = arrivalTimecountt
    sheet['C16'] = availCatCardcountt
    sheet['C17'] = availSrCitizencountt
    sheet['C18'] = availableSeatscountt
    sheet['C19'] = avlWindowSeatscountt
    sheet['C20'] = boardingTimescountt
    sheet['C21'] = bookablecountt
    sheet['C22'] = bpDpSeatLayoutcountt
    sheet['C23'] = busImageCountcountt
    sheet['C24'] = busServiceIdcountt
    sheet['C25'] = busTypecountt
    sheet['C26'] = busTypeIdcountt
    sheet['C27'] = cancellationPolicycountt
    sheet['C28'] = departureTimecountt
    sheet['C29'] = dojcountt
    sheet['C30'] = dropPointMandatorycountt
    sheet['C31'] = droppingTimescountt
    sheet['C32'] = fareDetailscountt
    sheet['C33'] = farescountt
    sheet['C34'] = flatComApplicablecountt
    sheet['C35'] = gdsCommissioncountt
    sheet['C36'] = idProofRequiredcountt
    sheet['C37'] = liveTrackingAvailablecountt
    sheet['C38'] = maxSeatsPerTicketcountt
    sheet['C39'] = nonACcountt
    sheet['C40'] = operatorcountt
    sheet['C41'] = otgEnabledcountt
    sheet['C42'] = otgPolicycountt
    sheet['C43'] = partialCancellationAllowedcountt
    sheet['C44'] = partnerBaseCommissioncountt
    sheet['C45'] = primaryPaxCancellablecountt
    sheet['C46'] = routeIdcountt
    sheet['C47'] = rtccountt
    sheet['C48'] = seatercountt
    sheet['C49'] = selfInventorycountt
    sheet['C50'] = singleLadiescountt
    sheet['C51'] = sleepercountt
    sheet['C52'] = tatkalTimecountt
    sheet['C53'] = vehicleTypecountt
    sheet['C54'] = viaRoutescountt
    sheet['C55'] = zeroCancellationTimecountt
    sheet['C56'] = mTicketEnabledcountt
    sheet['C57'] = sd_idcountt
    sheet['C58'] = createDtcountt
    sheet['C59'] = created_datecountt


    sheet['F9'] = (sourcecountt - sourcenamecountt)

    sheet['F11'] = (destinationcountt - destinationnamecountt)

    sheet['F13'] = 'travels - operators:' + str(travelscountt - operatorcountt)

   # listmin = [minidd, minsourcee, minsource_namee, mindestinationn, mindestinationnamee, mintravelss, minacc,
              # minarrivalTimee, minavailCatCardd, minavailSrCitizenn]

    # sheet['F18'] = "max seats 68?:" + '' + mymodule.val

    book.save(driver.desktop+'ParentMainAnalysis.xlsx')
    # book.save('DistinctCounts.xlsx')

    # except:
    # print("Something went wrong")

    conn.close()














